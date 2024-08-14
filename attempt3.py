import pandas as pd
import os
import sys
import pulp
from collections import defaultdict
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import colorsys


# Invigilator and Exam class Definitions
class Invigilator:
    def __init__(self, id, name, avail, lead, size_pref):
        self.id = int(id)
        self.name = name
        self.avail = [int(x) for x in avail.split(',')]
        self.lead = int(lead)
        self.size_pref = [str(x) for x in size_pref.split(',')]

    def __repr__(self):
        return f"{self.id},{self.name}, {self.avail}, {self.lead},{self.size_pref}"


class Exam:
    def __init__(self, id, name, date, session_size):
        self.id = int(id)
        self.name = name
        self.date = date
        self.session_size = int(session_size)
        self.invig_required, self.size_code = self.calculate_invig_requirements(session_size)

    def calculate_invig_requirements(self, session_size):
        if session_size == 1:
            return 1, 's'
        elif session_size <= 30:
            return 2, 's'
        elif session_size <= 80:
            return 3, 's'
        elif session_size <= 120:
            return 4, 'm'
        elif session_size <= 160:
            return 5, 'm'
        elif session_size <= 200:
            return 6, 'm'
        elif session_size <= 240:
            return 7, 'm'
        elif session_size <= 300:
            return 8, 'm'
        elif session_size <= 340:
            return 9, 'l'
        elif session_size <= 380:
            return 10, 'l'
        elif session_size <= 420:
            return 11, 'l'
        else:
            return 12, 'l'

    def __repr__(self):
        return f"\n{self.id}, {self.name}, {self.date},{self.session_size},{self.invig_required},{self.size_code}"


# Utility Functions

def get_resource_path(relative_path):
# Get the absolute path to the resource, for PyInstaller 
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, relative_path)


def read_invig_as_dict(filename):
    invigilators = []
    df = pd.read_excel(filename)
    for _, row in df.iterrows():
        invigilator = Invigilator(
            id=row['invig_id'],
            name=row['name'],
            avail=row['time_slot_availability'],
            lead=row['lead'],
            size_pref=row['size_preference']
        )
        invigilators.append(invigilator)
    return invigilators


def read_exams_as_dict(filename):
    sessions = defaultdict(list)
    df = pd.read_excel(filename)

    for _, row in df.iterrows():
        exam_id, exam_name, date = row['exam_id'], row['exam_name'], row['date']
        session_size, time_slot = row['session_size'], row['time_slot']

        if pd.isna(time_slot):
            print(f"Skipping exam {exam_name} with ID {exam_id} due to missing time_slot")
            continue

        try:
            time_slot = int(time_slot)
        except ValueError:
            print(f"Skipping exam {exam_name} with ID {exam_id} due to invalid time_slot value: {time_slot}")
            continue

        exam = Exam(exam_id, exam_name, date, session_size)
        sessions[time_slot].append(exam)

    return sessions


def import_files(invig_file, exams_file):
    invigilators = read_invig_as_dict(invig_file)
    exam_sessions = read_exams_as_dict(exams_file)
    return exam_sessions, invigilators


def print_unmet_requirements(unmet_invig_requirements, unmet_lead_requirements):
    # Print unmet requirements
    if unmet_invig_requirements or unmet_lead_requirements:
        print("\nExams with unmet requirements:")
        if unmet_invig_requirements:
            print("Unmet Invigilator Requirements:")
            for exam_name, time_slot in unmet_invig_requirements:
                print(f"Exam: {exam_name}, Time Slot: {time_slot}")
        if unmet_lead_requirements:
            print("Unmet Lead Requirements:")
            for exam_name, time_slot in unmet_lead_requirements:
                print(f"Exam: {exam_name}, Time Slot: {time_slot}")
    else:
        print("\nAll exam requirements have been met.")


def print_unavailable_assignments(unavailable_assignments):
    # Print invigilators assigned to unavailable timeslots
    if unavailable_assignments:
        print("\nInvigilators assigned to unavailable timeslots:")
        for invigilator_name, time_slot in unavailable_assignments:
            print(f"Invigilator: {invigilator_name}, Time Slot: {time_slot}")
    else:
        print("\nNo invigilators were assigned to unavailable timeslots.")


def generate_colour_palette(num_colours):
    colours = []
    for i in range(num_colours):
        hue = (i / num_colours)
        lightness = 0.6
        saturation = 0.8
        rgb = colorsys.hls_to_rgb(hue, lightness, saturation)
        hex_colour = "{:02x}{:02x}{:02x}".format(int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255))
        colours.append(hex_colour)
    return colours


def create_penalty_matrix(invigilators, exam_sessions):
    penalty_matrix = {
        's': {'s': 0, 'm': 10, 'l': 20},
        'm': {'s': 10, 'm': 0, 'l': 10},
        'l': {'s': 20, 'm': 10, 'l': 0}
    }

    penalty = {}
    for invigilator in invigilators:
        for time_slot in exam_sessions:
            for exam in exam_sessions[time_slot]:
                base_penalty = min(
                    [penalty_matrix[exam.size_code].get(pref, 20) for pref in invigilator.size_pref]
                )
                # Add a small random perturbation to the penalty to make the assignment fair
                penalty[(invigilator.name, time_slot, exam.id)] = base_penalty + random.uniform(0, 0.01)
    
    return penalty


def create_decision_variables(invigilators, exam_sessions):
    x = pulp.LpVariable.dicts("assign",
                              [(invigilator.name, time_slot, exam.id)
                               for invigilator in invigilators
                               for time_slot in exam_sessions
                               for exam in exam_sessions[time_slot]],
                              cat='Binary')

    y = pulp.LpVariable.dicts("unavailable",
                              [(invigilator.name, time_slot)
                               for invigilator in invigilators
                               for time_slot in exam_sessions
                               if time_slot not in invigilator.avail],
                              cat='Binary')

    unmet_invig = pulp.LpVariable.dicts("unmet",
                                        [(time_slot, exam.id)
                                         for time_slot in exam_sessions
                                         for exam in exam_sessions[time_slot]],
                                        cat='Binary')

    unmet_lead = pulp.LpVariable.dicts("unmet_lead",
                                       [(time_slot, exam.id)
                                        for time_slot in exam_sessions
                                        for exam in exam_sessions[time_slot]],
                                       cat='Binary')

    return x, y, unmet_invig, unmet_lead


def define_constraints(prob, invigilators, exam_sessions, x, y, unmet_invig, unmet_lead, max_time_slot):
    # Constraints for ensuring sufficient invigilators and lead examiner are assigned
    for time_slot, exams in exam_sessions.items():
        for exam in exams:
            # Ensure the required number of invigilators are assigned or mark as unmet
            prob += (pulp.lpSum([x[invigilator.name, time_slot, exam.id]
                                 for invigilator in invigilators])
                     + unmet_invig[time_slot, exam.id]) >= exam.invig_required, f"Session_{exam.id}_Requirement"
            # Ensure at least one lead examiner is assigned to each exam or mark as unmet
            prob += (pulp.lpSum([x[invigilator.name, time_slot, exam.id]
                                 for invigilator in invigilators
                                 if invigilator.lead == 1])
                     + unmet_lead[time_slot, exam.id]) >= 1, f"Session_{exam.id}_Lead_Requirement"
# Constraint to ensure no invigilator is assigned to two consecutive timeslots in the same day
    for invigilator in invigilators:
        for day_start in range(1, max_time_slot + 1, 3): # Iterate over the start of each day
            if day_start + 2 <= max_time_slot: # Ensure we do not go out of bounds
                # Check if there are exams in the timeslots before applying the constraints
                if day_start in exam_sessions and day_start + 1 in exam_sessions:
                    # No consecutive timeslot assignments for the first two timeslots in a day
                    prob += pulp.lpSum([x[invigilator.name, day_start, exam.id]
                                        for exam in exam_sessions[day_start]]) + \
                            pulp.lpSum([x[invigilator.name, day_start + 1, exam.id]
                                        for exam in exam_sessions[day_start + 1]]) <= 1, \
                        f"No_Consecutive_Assignments_{invigilator.name}_Day_{(day_start // 3) + 1}_first_pair"

                if day_start + 1 in exam_sessions and day_start + 2 in exam_sessions:
                    # No consecutive timeslot assignments for the second and third timeslots in a day
                    prob += pulp.lpSum([x[invigilator.name, day_start + 1, exam.id]
                                        for exam in exam_sessions[day_start + 1]]) + \
                            pulp.lpSum([x[invigilator.name, day_start + 2, exam.id]
                                        for exam in exam_sessions[day_start + 2]]) <= 1, \
                        f"No_Consecutive_Assignments_{invigilator.name}_Day_{(day_start // 3) + 1}_second_pair"
    # Constraint to ensure an invigilator is assigned to only one exam per time slot
    for invigilator in invigilators:
        for slot in range(1, max_time_slot + 1):
            prob += pulp.lpSum([x[invigilator.name, slot, exam.id]
                                for exam in exam_sessions[slot]]) <= 1, f"Invigilator_{invigilator.name}_Slot_{slot}"

    return prob

# Objective function: Minimize penalties and the number of unavailable assignments
def define_objective(prob, invigilators, exam_sessions, x, y, unmet_invig, unmet_lead, penalty):
    large_penalty = 1000 # Large penalty for unmet invigilation requirements
    unavailable_penalty = 100 # Penalty for assigning to unavailable time slots

    prob += (
        pulp.lpSum(
            [
                (x[invigilator.name, time_slot, exam.id] * penalty[(invigilator.name, time_slot, exam.id)])
                for invigilator in invigilators
                for time_slot in exam_sessions
                for exam in exam_sessions[time_slot]
            ]
        )
        + large_penalty * (
            pulp.lpSum(
                unmet_invig[time_slot, exam.id] for time_slot in exam_sessions for exam in exam_sessions[time_slot]
            )
            + pulp.lpSum(
                unmet_lead[time_slot, exam.id] for time_slot in exam_sessions for exam in exam_sessions[time_slot]
            )
        )
        + unavailable_penalty
        * pulp.lpSum(
            [
                y[invigilator.name, time_slot]
                for invigilator in invigilators
                for time_slot in exam_sessions
                if (invigilator.name, time_slot) in y
            ]
        ),
        "Total Penalty, Unmet Requirements, and Unavailable Assignments",
    )
    return prob


def solve_problem(prob):
    prob.solve()
    return pulp.LpStatus[prob.status], pulp.value(prob.objective)


def initialize_results(invigilators, max_time_slot):
    return {invigilator.name: {slot: [] for slot in range(1, max_time_slot + 1)} for invigilator in invigilators}

# Fill the results dictionary with exam names

def fill_results(results, invigilators, exam_sessions, x, used_invigilators):
    unavailable_assignments = []
    for invigilator in invigilators:
        for slot in range(1, max(exam_sessions.keys()) + 1):
            for exam in exam_sessions[slot]:
                if pulp.value(x[invigilator.name, slot, exam.id]) == 1:
                    results[invigilator.name][slot].append(exam.name)
                    used_invigilators.add(invigilator.name)
                    # Check if the slot is not in the invigilator's availability
                    if slot not in invigilator.avail:
                        unavailable_assignments.append((invigilator.name, slot))
    return results, used_invigilators, unavailable_assignments

# Identify exams with unmet requirements

def identify_unmet_requirements(exam_sessions, unmet_invig, unmet_lead):
    unmet_invig_requirements = []
    unmet_lead_requirements = []

    for time_slot in exam_sessions:
        for exam in exam_sessions[time_slot]:
            if pulp.value(unmet_invig[time_slot, exam.id]) == 1:
                unmet_invig_requirements.append((exam.name, time_slot))
            if pulp.value(unmet_lead[time_slot, exam.id]) == 1:
                unmet_lead_requirements.append((exam.name, time_slot))

    return unmet_invig_requirements, unmet_lead_requirements


def export_results_to_excel(invigilators, exam_sessions, results, exam_colours):
    max_time_slot = max(exam_sessions.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Invigilator Assignments"

    columns = ["Invigilator"] + [f"Slot {slot}" for slot in range(1, max_time_slot + 1)]
    ws.append(columns)

    for invigilator in sorted(invigilators, key=lambda inv: inv.name):
        row = [invigilator.name]
        for slot in range(1, max_time_slot + 1):
            exams_for_slot = results[invigilator.name][slot]
            row.append(",".join(exams_for_slot) if exams_for_slot else "-")
        ws.append(row)

    format_cells(ws, invigilators, exam_colours)
    wb.save('invigilator_assignments.xlsx')
    print("Results have been exported to invigilator_assignments.xlsx")

# Colours the font and cells of each assignment, white font for leads, black for no leads, each exam has it own colour for readability
def format_cells(ws, invigilators, exam_colours):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value and cell.value != "-":
                exam_ids_in_cell = extract_exam_ids_from_cell(cell.value)
                invigilator_name = ws.cell(row=cell.row, column=1).value
                invigilator = next((inv for inv in invigilators if inv.name == invigilator_name), None)
                if invigilator and invigilator.lead == 1:
                    cell.font = Font(color="FFFFFF")
                apply_cell_format(cell, exam_ids_in_cell, exam_colours)


def extract_exam_ids_from_cell(cell_value):
    exam_ids = []
    for exam_id in cell_value.split(","):
        numeric_id = ''.join(filter(str.isdigit, exam_id))
        if numeric_id:
            exam_ids.append(int(numeric_id))
    return exam_ids


def apply_cell_format(cell, exam_ids_in_cell, exam_colours):
    for exam_id in exam_ids_in_cell:
        if exam_id in exam_colours:
            cell.fill = PatternFill(start_color=exam_colours[exam_id], end_color=exam_colours[exam_id], fill_type="solid")


def print_unassigned_invigilators(invigilators, results, max_time_slot):
    # Identify invigilators with no assignments
    unassigned_invigilators = [invigilator.name for invigilator in invigilators if not any(results[invigilator.name][slot] for slot in range(1, max_time_slot + 1))]

    # Print the list of unassigned invigilators
    if unassigned_invigilators:
        print("\nInvigilators with no assignments:")
        for invigilator_name in sorted(unassigned_invigilators):
            print(f"Invigilator: {invigilator_name}")
    else:
        print("\nAll invigilators have assignments.")


def main():
    # Paths to your Excel files
    invig_file = get_resource_path('large_invigilators.xlsx')
    exams_file = get_resource_path('large_exam_venues.xlsx')

    exam_sessions, invigilators = import_files(invig_file, exams_file)

    max_time_slot = max(exam_sessions.keys())
    random.shuffle(invigilators)
    x, y, unmet_invig, unmet_lead = create_decision_variables(invigilators, exam_sessions)

    prob = pulp.LpProblem("Invigilator_Allocation", pulp.LpMinimize)

    penalty = create_penalty_matrix(invigilators, exam_sessions)

    prob = define_constraints(prob, invigilators, exam_sessions, x, y, unmet_invig, unmet_lead, max_time_slot)

    prob = define_objective(prob, invigilators, exam_sessions, x, y, unmet_invig, unmet_lead, penalty)

    status, objective_value = solve_problem(prob)

    print(f"Status: {status}")
    print(f"Total Penalty: {objective_value}")

    results = initialize_results(invigilators, max_time_slot)

    used_invigilators = set()
    results, used_invigilators, unavailable_assignments = fill_results(results, invigilators, exam_sessions, x, used_invigilators)

    unmet_invig_requirements, unmet_lead_requirements = identify_unmet_requirements(exam_sessions, unmet_invig, unmet_lead)

    exam_ids = [exam.id for time_slot in exam_sessions for exam in exam_sessions[time_slot]]
    exam_colours = generate_colour_palette(len(exam_ids) + 10)

    exam_colours_mapping = {}
    for i in range(len(exam_ids) + 1):
        colour_index = (i * 6) % len(exam_colours)
        exam_colours_mapping[i] = exam_colours[colour_index]

    export_results_to_excel(invigilators, exam_sessions, results, exam_colours_mapping)

    # Print unmet requirements
    print_unmet_requirements(unmet_invig_requirements, unmet_lead_requirements)

    # Print unassigned invigilators
    print_unassigned_invigilators(invigilators, results, max_time_slot)

    # Print invigilators assigned to unavailable timeslots
    print_unavailable_assignments(unavailable_assignments)


if __name__ == "__main__":
    main()
