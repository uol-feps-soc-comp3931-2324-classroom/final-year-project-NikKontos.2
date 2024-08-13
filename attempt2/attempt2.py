import pandas as pd
import pulp
from collections import defaultdict
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import colorsys

class Invigilator:
    def __init__(self, id, name, avail, lead, size_pref):
        self.id = int(id)
        self.name = name
        self.avail = [int(x) for x in avail.split(',')]
        self.lead = int(lead)
        self.size_pref = [str(x) for x in size_pref.split(',')]
    def __repr__(self):
        return f"{self.id},{self.name}, {self.avail}, {self.lead},{self.size_pref}"

class Exam:
    def __init__(self, id, name, date, session_size):
        self.id = int(id)
        self.name = name
        self.date = date
        self.session_size = int(session_size)
        
        if self.session_size == 1:
            invig_required = 1
            size_code = 's'
        elif self.session_size <= 30:
            invig_required = 2
            size_code = 's'
        elif self.session_size <= 80:
            invig_required = 3
            size_code = 's'
        elif self.session_size <= 120:
            invig_required = 4
            size_code = 'm'
        elif self.session_size <= 160:
            invig_required = 5
            size_code = 'm'
        elif self.session_size <= 200:
            invig_required = 6
            size_code = 'm'
        elif self.session_size <= 240:
            invig_required = 7
            size_code = 'm'
        elif self.session_size <= 300:
            invig_required = 8
            size_code = 'm'
        elif self.session_size <= 340:
            invig_required = 9
            size_code = 'l'
        elif self.session_size <= 380:
            invig_required = 10
            size_code = 'l'
        elif self.session_size <= 420:
            invig_required = 11
            size_code = 'l'
        else:
            invig_required = 12
            size_code = 'l'

        self.invig_required = int(invig_required)
        self.size_code = size_code
    def __repr__(self):
        return f"\n{self.id}, {self.name}, {self.date},{self.session_size},{self.invig_required},{self.size_code}"

def read_invig_as_dict(filename):
    invigilators = []
    df = pd.read_excel(filename)
    for _, row in df.iterrows():
        invig_id, invig_name = row['invig_id'], row['name']
        avail, lead, size_pref = row['time_slot_availability'], row['lead'], row['size_preference']
        invigilator = Invigilator(invig_id, invig_name, avail, lead, size_pref)
        invigilators.append(invigilator)
    random.shuffle(invigilators)
    return invigilators

def read_exams_as_dict(filename):
    sessions = defaultdict(list)
    df = pd.read_excel(filename)
    for _, row in df.iterrows():
        exam_id, exam_name, date = row['exam_id'], row['exam_name'], row['date']
        session_size, time_slot = row['session_size'], row['time_slot']
        
        # Check if time_slot is NaN or cannot be converted to an integer
        if pd.isna(time_slot):
            print(f"Skipping exam {exam_name} with ID {exam_id} due to missing time_slot")
            continue
        
        try:
            time_slot = int(time_slot)
        except ValueError:
            print(f"Skipping exam {exam_name} with ID {exam_id} due to invalid time_slot value: {time_slot}")
            continue
        
        exam = Exam(exam_id, exam_name, date, session_size)
        sessions[time_slot].append(exam)
    
    return sessions

def import_files(invig_file, exams_file):
    invigilators = read_invig_as_dict(invig_file)
    exam_sessions = read_exams_as_dict(exams_file)
    return exam_sessions, invigilators

def output_imports():
    print(invigilators, exam_sessions)

invig_file = 'large_invigilators.xlsx'
exams_file = 'large_exam_venues.xlsx'

exam_sessions, invigilators = import_files(invig_file, exams_file)

output_imports()

max_time_slot = max(exam_sessions.keys())

# Create the problem variable
prob = pulp.LpProblem("Invigilator_Allocation", pulp.LpMinimize)

# Decision Variables
x = pulp.LpVariable.dicts("assign", 
                         [(invigilator.name, time_slot, exam.id) 
                          for invigilator in invigilators 
                          for time_slot in exam_sessions
                          for exam in exam_sessions[time_slot]], 
                         cat='Binary')

# Variable to track if an invigilator is assigned to an unavailable timeslot
y = pulp.LpVariable.dicts("unavailable", 
                         [(invigilator.name, time_slot) 
                          for invigilator in invigilators 
                          for time_slot in exam_sessions], 
                         cat='Binary')

# Add constraints to track unavailable assignments
for invigilator in invigilators:
    for time_slot in exam_sessions:
        if time_slot not in invigilator.avail:
            prob += pulp.lpSum([x[invigilator.name, time_slot, exam.id] 
                                for exam in exam_sessions[time_slot]]) <= y[invigilator.name, time_slot] * len(exam_sessions[time_slot]), \
                                f"Unavailable_Assignment_{invigilator.name}_{time_slot}"

# Variable for unmet invigilation requirements
unmet_invig = pulp.LpVariable.dicts("unmet", 
                                    [(time_slot, exam.id) 
                                     for time_slot in exam_sessions
                                     for exam in exam_sessions[time_slot]], 
                                    cat='Binary')

# Variable for unmet lead examiner requirement
unmet_lead = pulp.LpVariable.dicts("unmet_lead", 
                                    [(time_slot, exam.id) 
                                     for time_slot in exam_sessions
                                     for exam in exam_sessions[time_slot]], 
                                    cat='Binary')

# Penalties
penalty_matrix = {
    's': {'s': 0, 'm': 10, 'l': 20},
    'm': {'s': 10, 'm': 0, 'l': 10},
    'l': {'s': 20, 'm': 10, 'l': 0}
}

penalty = {}
for invigilator in invigilators:
    for time_slot in exam_sessions:
        for exam in exam_sessions[time_slot]:
            penalty[(invigilator.name, time_slot, exam.id)] = min(
                [penalty_matrix[exam.size_code].get(pref, 20) for pref in invigilator.size_pref]
            )

# Objective function: Minimize penalties and the number of unavailable assignments
large_penalty = 1000  # Large penalty for unmet invigilation requirements
unavailable_penalty = 100  # Penalty for assigning to unavailable time slots

prob += (
    pulp.lpSum(
        [
            (x[invigilator.name, time_slot, exam.id] * penalty[(invigilator.name, time_slot, exam.id)]
             + random.uniform(0, 1e-5))  # Add small random perturbation
            for invigilator in invigilators
            for time_slot in exam_sessions
            for exam in exam_sessions[time_slot]
        ]
    )
    + large_penalty * (
        pulp.lpSum(
            unmet_invig[time_slot, exam.id] for time_slot in exam_sessions for exam in exam_sessions[time_slot]
        )
        + pulp.lpSum(
            unmet_lead[time_slot, exam.id] for time_slot in exam_sessions for exam in exam_sessions[time_slot]
        )
    )
    + unavailable_penalty
    * pulp.lpSum(
        [
            y[invigilator.name, time_slot]
            for invigilator in invigilators
            for time_slot in exam_sessions
        ]
    ),
    "Total Penalty, Unmet Requirements, and Unavailable Assignments",
)


# Constraints for ensuring sufficient invigilators and lead examiner are assigned
for time_slot, exams in exam_sessions.items():
    for exam in exams:
        # Ensure required number of invigilators are assigned or mark as unmet
        prob += (pulp.lpSum([x[invigilator.name, time_slot, exam.id] 
                             for invigilator in invigilators]) 
                + unmet_invig[time_slot, exam.id]) >= exam.invig_required, f"Session_{exam.id}_Requirement"
        
        # Ensure at least one lead examiner is assigned to each exam or mark as unmet
        prob += (pulp.lpSum([x[invigilator.name, time_slot, exam.id] 
                            for invigilator in invigilators 
                            if invigilator.lead == 1]) 
                + unmet_lead[time_slot, exam.id]) >= 1, f"Session_{exam.id}_Lead_Requirement"
# Constraint to ensure no invigilator is assigned to two consecutive timeslots in the same day
for invigilator in invigilators:
    for day_start in range(1, max_time_slot + 1, 3):  # Iterate over the start of each day
        if day_start + 2 <= max_time_slot:  # Ensure we do not go out of bounds
            # Check if there are exams in the timeslots before applying the constraints
            if day_start in exam_sessions and day_start + 1 in exam_sessions:
                # No consecutive timeslot assignments for the first two timeslots in a day
                prob += pulp.lpSum([x[invigilator.name, day_start, exam.id] 
                                    for exam in exam_sessions[day_start]]) + \
                        pulp.lpSum([x[invigilator.name, day_start + 1, exam.id] 
                                    for exam in exam_sessions[day_start + 1]]) <= 1, \
                        f"No_Consecutive_Assignments_{invigilator.name}_Day_{(day_start // 3) + 1}_first_pair"

            if day_start + 1 in exam_sessions and day_start + 2 in exam_sessions:
                # No consecutive timeslot assignments for the second and third timeslots in a day
                prob += pulp.lpSum([x[invigilator.name, day_start + 1, exam.id] 
                                    for exam in exam_sessions[day_start + 1]]) + \
                        pulp.lpSum([x[invigilator.name, day_start + 2, exam.id] 
                                    for exam in exam_sessions[day_start + 2]]) <= 1, \
                        f"No_Consecutive_Assignments_{invigilator.name}_Day_{(day_start // 3) + 1}_second_pair"

# Constraint to ensure an invigilator is assigned to only one exam per time slot
for invigilator in invigilators:
    for slot in range(1, max_time_slot+1):  # Iterate over all possible slots
        prob += pulp.lpSum([x[invigilator.name, slot, exam.id] 
                            for exam in exam_sessions[slot]]) <= 1, f"Invigilator_{invigilator.name}_Slot_{slot}"

# Solve the problem
prob.solve()

# Initialize results dictionary with all slots
results = {invigilator.name: {slot: [] for slot in range(1, max_time_slot+1)} for invigilator in invigilators}

# Track used invigilators
used_invigilators = set()

# Track unmet requirements
unmet_invig_requirements = []
unmet_lead_requirements = []

# Fill the results dictionary with exam names
for invigilator in invigilators:
    for slot in range(1, max_time_slot+1):  # Iterate over all possible slots
        for exam in exam_sessions[slot]:
            if pulp.value(x[invigilator.name, slot, exam.id]) == 1:
                results[invigilator.name][slot].append(exam.name)
                used_invigilators.add(invigilator.name)

# Identify exams with unmet requirements
for time_slot in exam_sessions:
    for exam in exam_sessions[time_slot]:
        if pulp.value(unmet_invig[time_slot, exam.id]) == 1:
            unmet_invig_requirements.append((exam.name, time_slot))
        if pulp.value(unmet_lead[time_slot, exam.id]) == 1:
            unmet_lead_requirements.append((exam.name, time_slot))

# Print results
print("Invigilator\\TimeSlot\t" + "\t".join(map(str, range(1, max_time_slot+1))))
for invigilator in invigilators:
    row = f"{invigilator.name}"
    for slot in range(1, max_time_slot+1):
        exams_for_slot = results[invigilator.name][slot]
        if exams_for_slot:
            row += "\t" + ",".join(exams_for_slot)
        else:
            row += "\t-"
    print(row)

print(f"Status: {pulp.LpStatus[prob.status]}")
print(f"Total Invigilators Used: {len(used_invigilators)}")
print(f"Total Penalty: {pulp.value(prob.objective)}")

# Print unmet requirements
if unmet_invig_requirements or unmet_lead_requirements:
    print("\nExams with unmet requirements:")
    if unmet_invig_requirements:
        print("Unmet Invigilator Requirements:")
        for exam_name, time_slot in unmet_invig_requirements:
            print(f"Exam: {exam_name}, Time Slot: {time_slot}")
    if unmet_lead_requirements:
        print("Unmet Lead Requirements:")
        for exam_name, time_slot in unmet_lead_requirements:
            print(f"Exam: {exam_name}, Time Slot: {time_slot}")
else:
    print("\nAll exam requirements have been met.")

# List invigilators assigned to unavailable slots
unavailable_assignments = []

for invigilator in invigilators:
    for time_slot in exam_sessions:
        if (invigilator.name, time_slot) in y:  # Check if this time slot exists for this invigilator
            if pulp.value(y[invigilator.name, time_slot]) == 1:
                unavailable_assignments.append((invigilator.name, time_slot))

# Sort the list of unavailable assignments alphabetically by invigilator name
unavailable_assignments.sort()

# Print the list of invigilators assigned to unavailable slots
if unavailable_assignments:
    print("\nInvigilators assigned to unavailable slots:")
    for invigilator_name, time_slot in unavailable_assignments:
        print(f"Invigilator: {invigilator_name}, Time Slot: {time_slot}")
else:
    print("\nNo invigilators were assigned to unavailable slots.")

# Identify invigilators with no assignments
unassigned_invigilators = [invigilator.name for invigilator in invigilators if not any(results[invigilator.name][slot] for slot in range(1, max_time_slot+1))]

# Print the list of unassigned invigilators
if unassigned_invigilators:
    print("\nInvigilators with no assignments:")
    for invigilator_name in sorted(unassigned_invigilators):
        print(f"Invigilator: {invigilator_name}")
else:
    print("\nAll invigilators have assignments.")


# Generate colours that change progressively and ensure readability
def generate_colour_palette(num_colours):
    colours = []
    for i in range(num_colours):
        hue = (i / num_colours)
        lightness=0.6 # Reduce lightness to avoid too light colours 
        saturation = 0.8  # Increase saturation for more vivid colours 
        rgb = colorsys.hls_to_rgb(hue, lightness, saturation)
        hex_colour = "{:02x}{:02x}{:02x}".format(int(rgb[0]*255), int(rgb[1]*255), int(rgb[2]*255))
        colours.append(hex_colour)
    print("HEX__",hex_colour)
    return colours
# Ensure exam_ids are consistently tracked and associated with colours
exam_ids = [exam.id for time_slot in exam_sessions for exam in exam_sessions[time_slot]]

# Generate and assign colours

exam_colours = {}
colour_palette = generate_colour_palette(len(exam_ids) + 10)  # Ensure the palette is large enough

for  i in range(len(exam_ids)+1):
    # Use modular arithmetic to skip 5 colours for each exam
    colour_index = (i * 6) % len(colour_palette)
    exam_colours[i] = colour_palette[colour_index]
# Prepare data for Excel output
columns = ["Invigilator"] + [f"Slot {slot}" for slot in range(1, max_time_slot + 1)]
wb = Workbook()
ws = wb.active
ws.title = "Invigilator Assignments"

# Write the header
ws.append(columns)

# Write results to the sheet
for invigilator in sorted(invigilators, key=lambda inv: inv.name):
    row = [invigilator.name]
    for slot in range(1, max_time_slot + 1):
        exams_for_slot = results[invigilator.name][slot]
        if exams_for_slot:
            exam_list = ",".join(exams_for_slot)
            row.append(exam_list)
        else:
            row.append("-")
    ws.append(row)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
    for cell in row:
        if cell.value and cell.value != "-":
            # Extract exam IDs, handling cases where non-numeric values might be present
            exam_ids_in_cell = []
            for exam_id in cell.value.split(","):
                try:
                    # Extract numeric part from possible mixed strings
                    numeric_id = ''.join(filter(str.isdigit, exam_id))
                    if numeric_id:
                        exam_ids_in_cell.append(int(numeric_id))
                except ValueError:
                    print(f"Skipping invalid exam ID: {exam_id}")
            
            invigilator_name = ws.cell(row=cell.row, column=1).value
            invigilator = next((inv for inv in invigilators if inv.name == invigilator_name), None)
            if invigilator and invigilator.lead == 1:
                cell.font = Font(color="FFFFFF")
            for exam_id in exam_ids_in_cell:
                if exam_id in exam_colours:
                    cell.fill = PatternFill(start_color=exam_colours[exam_id], end_color=exam_colours[exam_id], fill_type="solid")
print(exam_colours)
            
# Save the workbook
wb.save('invigilator_assignments.xlsx')

print("Results have been exported to invigilator_assignments.xlsx\nEach exam is assigned a specific colour.\nLead examiners assignments are displayed in white text.")