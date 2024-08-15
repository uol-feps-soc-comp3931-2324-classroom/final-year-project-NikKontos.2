
import subprocess
import sys
import pkg_resources
import datetime
import math
import pandas as pd
import os
import pulp
import random
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import colorsys

# Installation and Version Checking
def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

def uninstall(package):
    subprocess.check_call([sys.executable, "-m", "pip", "uninstall", "-y", package])

def check_and_install():
    # Define the required versions
    required_packages = {
        "numpy": "1.21.6",
        "pandas": "1.3.5",
        "pulp": "2.9.0",
        "openpyxl": "3.1.5"
    }

    for package, required_version in required_packages.items():
        try:
            installed_version = pkg_resources.get_distribution(package).version
            if installed_version != required_version:
                print(f"{package} version {installed_version} found, but {required_version} is required. Reinstalling...")
                uninstall(package)
                install(f"{package}=={required_version}")
            else:
                print(f"{package} {required_version} is already installed.")
        except pkg_resources.DistributionNotFound:
            print(f"{package} is not installed. Installing {required_version}...")
            install(f"{package}=={required_version}")

check_and_install()

# Class Definitions
class Invigilator:
    def __init__(self, id, name, avail, lead, size_pref):
        self.id = int(id)
        self.name = name
        self.avail = [int(x) for x in avail.split(',')]
        self.lead = int(lead)
        self.size_pref = [str(x) for x in size_pref.split(',')]

    def __repr__(self):
        return f"{self.id},{self.name}, {self.avail}, {self.lead},{self.size_pref}"

class Exam:
    def __init__(self, id, name, date, session_size):
        self.id = int(id)
        self.name = name
        self.date = date
        self.session_size = int(session_size)
        self.invig_required, self.size_code = self.calculate_invig_requirements(session_size)

    def calculate_invig_requirements(self, session_size):
        if session_size == 1:
            return 1, 's'
        elif session_size <= 30:
            return 2, 's'
        elif session_size <= 80:
            return 3, 's'
        elif session_size <= 120:
            return 4, 'm'
        elif session_size <= 160:
            return 5, 'm'
        elif session_size <= 200:
            return 6, 'm'
        elif session_size <= 240:
            return 7, 'm'
        elif session_size <= 300:
            return 8, 'm'
        elif session_size <= 340:
            return 9, 'l'
        elif session_size <= 380:
            return 10, 'l'
        elif session_size <= 420:
            return 11, 'l'
        else:
            return 12, 'l'

    def __repr__(self):
        return f"\n{self.id}, {self.name}, {self.date},{self.session_size},{self.invig_required},{self.size_code}"

# Utility Functions
def get_resource_path(relative_path):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, relative_path)

def read_invig_as_dict(filename):
    invigilators = []
    df = pd.read_excel(filename)
    for _, row in df.iterrows():
        invigilator = Invigilator(
            id=row['invig_id'],
            name=row['name'],
            avail=row['time_slot_availability'],
            lead=row['lead'],
            size_pref=row['size_preference']
        )
        invigilators.append(invigilator)
    return invigilators

def read_exams_as_dict(filename):
    sessions = defaultdict(list)
    df = pd.read_excel(filename)

    for _, row in df.iterrows():
        exam_id, exam_name, date = row['exam_id'], row['exam_name'], row['date']
        session_size, time_slot = row['session_size'], row['time_slot']

        if pd.isna(time_slot):
            print(f"Skipping exam {exam_name} with ID {exam_id} due to missing time_slot")
            continue

        try:
            time_slot = int(time_slot)
        except ValueError:
            print(f"Skipping exam {exam_name} with ID {exam_id} due to invalid time_slot value: {time_slot}")
            continue

        exam = Exam(exam_id, exam_name, date, session_size)
        sessions[time_slot].append(exam)

    return sessions

def import_files(invig_file, exams_file):
    invigilators = read_invig_as_dict(invig_file)
    exam_sessions = read_exams_as_dict(exams_file)
    return exam_sessions, invigilators

def generate_colour_palette(num_colours):
    colours = []
    for i in range (num_colours):
        hue = (i / num_colours)
        lightness = 0.6
        saturation = 0.8
        rgb = colorsys.hls_to_rgb(hue, lightness, saturation)
        hex_colour = "{:02x}{:02x}{:02x}".format(int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255))
        colours.append(hex_colour)
    return colours

# Penalty Matrix and Decision Variables
def create_penalty_matrix(invigilators, exam_sessions):
    penalty_matrix = {
        's': {'s': 0, 'm': 10, 'l': 20},
        'm': {'s': 10, 'm': 0, 'l': 10},
        'l': {'s': 20, 'm': 10, 'l': 0}
    }

    penalty = {}
    for invigilator in invigilators:
        for time_slot in exam_sessions:
            for exam in exam_sessions[time_slot]:
                base_penalty = min(
                    [penalty_matrix[exam.size_code].get(pref, 20) for pref in invigilator.size_pref]
                )
                penalty[(invigilator.name, time_slot, exam.id)] = base_penalty + random.uniform(0, 0.01)
    
    return penalty

def create_decision_variables(invigilators, exam_sessions):
    x = pulp.LpVariable.dicts("assign",
                              [(invigilator.name, time_slot, exam.id)
                               for invigilator in invigilators
                               for time_slot in exam_sessions
                               for exam in exam_sessions[time_slot]],
                              cat='Binary')

    y = pulp.LpVariable.dicts("unavailable",
                              [(invigilator.name, time_slot)
                               for invigilator in invigilators
                               for time_slot in exam_sessions
                               if time_slot not in invigilator.avail],
                              cat='Binary')

    unmet_invig = pulp.LpVariable.dicts("unmet",
                                        [(time_slot, exam.id)
                                         for time_slot in exam_sessions
                                         for exam in exam_sessions[time_slot]],
                                        cat='Binary')

    unmet_lead = pulp.LpVariable.dicts("unmet_lead",
                                       [(time_slot, exam.id)
                                        for time_slot in exam_sessions
                                        for exam in exam_sessions[time_slot]],
                                       cat='Binary')

    return x, y, unmet_invig, unmet_lead

# Constraint and Objective Definitions
def define_constraints(prob, invigilators, exam_sessions, x, y, unmet_invig, unmet_lead, max_time_slot):
    # Ensure sufficient invigilators and lead examiners
    for time_slot, exams in exam_sessions.items():
        for exam in exams:
            prob += (pulp.lpSum([x[invigilator.name, time_slot, exam.id] for invigilator in invigilators])
                     + unmet_invig[time_slot, exam.id]) >= exam.invig_required, f"Session_{exam.id}_Requirement"
            prob += (pulp.lpSum([x[invigilator.name, time_slot, exam.id] for invigilator in invigilators if invigilator.lead == 1])
                     + unmet_lead[time_slot, exam.id]) >= 1, f"Session_{exam.id}_Lead_Requirement"

    # Ensure no invigilator is assigned to two consecutive timeslots in the same day
    for invigilator in invigilators:
        for day_start in range (1, max_time_slot + 1, 3):
            if day_start + 2 <= max_time_slot:
                if day_start in exam_sessions and day_start + 1 in exam_sessions:
                    prob += pulp.lpSum([x[invigilator.name, day_start, exam.id] for exam in exam_sessions[day_start]]) + \
                            pulp.lpSum([x[invigilator.name, day_start + 1, exam.id] for exam in exam_sessions[day_start + 1]]) <= 1, \
                            f"No_Consecutive_Assignments_{invigilator.name}_Day_{(day_start // 3) + 1}_first_pair"

                if day_start + 1 in exam_sessions and day_start + 2 in exam_sessions:
                    prob += pulp.lpSum([x[invigilator.name, day_start + 1, exam.id] for exam in exam_sessions[day_start + 1]]) + \
                            pulp.lpSum([x[invigilator.name, day_start + 2, exam.id] for exam in exam_sessions[day_start + 2]]) <= 1, \
                            f"No_Consecutive_Assignments_{invigilator.name}_Day_{(day_start // 3) + 1}_second_pair"

     # Ensure an invigilator is assigned to only one exam per time slot
    for invigilator in invigilators:
        for slot in range (1, max_time_slot + 1):
            prob += pulp.lpSum([x[invigilator.name, slot, exam.id] for exam in exam_sessions[slot]]) <= 1, \
                    f"Invigilator_{invigilator.name}_Slot_{slot}"

    return prob

def define_objective(prob, invigilators, exam_sessions, x, y, unmet_invig, unmet_lead, penalty):
    large_penalty = 1000  # Large penalty for unmet invigilation requirements
    unavailable_penalty = 100  # Penalty for assigning to unavailable time slots

    prob += (
        pulp.lpSum(
            [
                (x[invigilator.name, time_slot, exam.id] * penalty[(invigilator.name, time_slot, exam.id)])
                for invigilator in invigilators
                for time_slot in exam_sessions
                for exam in exam_sessions[time_slot]
            ]
        )
        + large_penalty * (
            pulp.lpSum(
                unmet_invig[time_slot, exam.id] for time_slot in exam_sessions for exam in exam_sessions[time_slot]
            )
            + pulp.lpSum(
                unmet_lead[time_slot, exam.id] for time_slot in exam_sessions for exam in exam_sessions[time_slot]
            )
        )
        + unavailable_penalty * pulp.lpSum(
            y[invigilator.name, time_slot] for invigilator in invigilators for time_slot in exam_sessions if (invigilator.name, time_slot) in y
        ),
        "Total Penalty, Unmet Requirements, and Unavailable Assignments"
    )
    return prob

 # Solution Handling
def solve_problem(prob):
    prob.solve()
    return pulp.LpStatus[prob.status], pulp.value(prob.objective)

def initialize_results(invigilators, max_time_slot):
    return {invigilator.name: {slot: [] for slot in range(1, max_time_slot + 1)} for invigilator in invigilators}

def fill_results(results, invigilators, exam_sessions, x, used_invigilators):
    unavailable_assignments =  []
    for invigilator in invigilators:
        for slot in range(1, max(exam_sessions.keys()) + 1):
            for exam in exam_sessions[slot]:
                if pulp.value(x[invigilator.name, slot, exam.id]) == 1:
                    results[invigilator.name][slot].append(exam.name)
                    used_invigilators.add(invigilator.name)
                    if slot not in invigilator.avail:
                        unavailable_assignments.append((invigilator.name, slot, exam.name))
    return results, used_invigilators, unavailable_assignments

def identify_unmet_requirements(exam_sessions, unmet_invig, unmet_lead):
    unmet_invig_requirements = []
    unmet_lead_requirements = []

    for time_slot in exam_sessions:
        for exam in exam_sessions[time_slot]:
            if pulp.value(unmet_invig[time_slot, exam.id]) == 1:
                unmet_invig_requirements.append((exam.name, time_slot))
            if pulp.value(unmet_lead[time_slot, exam.id]) == 1:
                unmet_lead_requirements.append((exam.name, time_slot))

    return unmet_invig_requirements, unmet_lead_requirements

# Excel Export and Formatting
def export_results_to_excel(invigilators, exam_sessions, results, exam_colours, unmet_invig_requirements, unmet_lead_requirements, unavailable_assignments, unassigned_invigilators):
    max_time_slot = max(exam_sessions.keys() )

    wb = Workbook()
    ws = wb.active
    ws.title = "Invigilator Assignments"

    # Add invigilator assignment data
    columns = ["Invigilator"] + [f"Slot {slot}" for slot in range(1, max_time_slot + 1)]
    ws.append(columns)

    for invigilator in sorted(invigilators, key=lambda inv: inv.name):
        row = [invigilator.name]
        for slot in range(1, max_time_slot + 1):
            exams_for_slot = results[invigilator.name][slot]
            row.append(",".join(exams_for_slot) if exams_for_slot else "-")
        ws.append(row)

    # Add unmet lead requirements
    start_col = len(columns) + 4
    start_col_letter = get_column_letter(start_col)
    ws[f"{start_col_letter}1"] = "Unmet Lead Requirements"
    
    row_offset = 2
    for exam_name, time_slot in unmet_lead_requirements:
        ws[f"{start_col_letter}{row_offset}"] = f"Exam: {exam_name}, Time Slot: {time_slot}"
        row_offset += 1

    # Add invigilators assigned to unavailable timeslots
    start_col = len(columns) + 5
    start_col_letter =  get_column_letter(start_col)
    ws[f"{start_col_letter}1"] = "Invigilators Assigned to Unavailable Timeslots"
    
    row_offset = 2
    for invigilator_name, time_slot, exam_name in unavailable_assignments:
        ws[f"{start_col_letter}{row_offset}"] = f"Invigilator: {invigilator_name}, Time Slot: {time_slot}, Exam: {exam_name}"
        row_offset += 1

    # Add unassigned invigilators
    start_col = len(columns) + 6
    start_col_letter = get_column_letter(start_col)
    ws[f"{start_col_letter}1"] = "Unassigned Invigilators"
    
    row_offset = 2
    for invigilator_name in unassigned_invigilators:
        ws[f"{start_col_letter}{row_offset}"] = f"Invigilator: {invigilator_name}"
        row_offset += 1

    # Format cells and auto-fit columns
    format_cells(ws, invigilators, exam_colours, exam_sessions)

    # Auto-size columns based on content
    auto_size_columns(ws)

    # Save the workbook with a timestamp
    now = datetime.datetime.now()
    timestamp = now.strftime("%d.%m.%Y_(%H,%M,%S)")
    filename = f'invigilator_assignments_{timestamp}.xlsx'
    wb.save(filename)
    print(f"Results have been exported to {filename}")

def auto_size_columns(ws, padding=2):
    
   #Auto-sizes the columns in the worksheet based on the length of the content,
    #adding a fixed padding to ensure readability.

    
    #padding: The number of extra spaces to add to the column width.

    for col in ws.columns:
        max_length = 0
        col_letter  = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        # Adjust width by adding a fixed padding
        adjusted_width = max_length + padding
        ws.column_dimensions[col_letter].width = adjusted_width


def format_cells(ws, invigilators, exam_colours, exam_sessions):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value and cell.value != "-":
                exam_ids_in_cell = extract_exam_ids_from_cell(cell.value)
                invigilator_name = ws.cell(row=cell.row, column=1).value
                invigilator = next((inv for inv in invigilators if inv.name == invigilator_name), None)
                if invigilator and invigilator.lead == 1:
                    cell.font = Font(color="FFFFFF")
                apply_cell_format(cell, exam_ids_in_cell, exam_colours)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.max_column - 5, max_col=ws.max_column - 5):
        for cell in row:
            if cell.value and "Exam:" in cell.value:
                exam_name = cell.value.split(",")[0].split(":")[1].strip()
                exam_id = get_exam_id_from_name(exam_name, exam_sessions)
                if exam_id in exam_colours:
                    cell.fill = PatternFill(start_color=exam_colours[exam_id], end_color=exam_colours[exam_id], fill_type="solid")

def extract_exam_ids_from_cell(cell_value):
    exam_ids = []
    for exam_id in cell_value.split(","):
        numeric_id = ''.join(filter(str.isdigit, exam_id))
        if numeric_id:
            exam_ids.append(int(numeric_id))
    return exam_ids

def get_exam_id_from_name(exam_name, exam_sessions):
    for time_slot, exams in exam_sessions.items():
        for exam in exams:
            if exam.name == exam_name:
                return exam.id
    return None

def apply_cell_format(cell, exam_ids_in_cell, exam_colours):
    for exam_id in exam_ids_in_cell:
        if exam_id in exam_colours:
            cell.fill = PatternFill(start_color=exam_colours[exam_id], end_color=exam_colours[exam_id], fill_type="solid")

# Utility Functions for Printing
def print_unmet_requirements(unmet_invig_requirements, unmet_lead_requirements):
    if unmet_invig_requirements or unmet_lead_requirements:
        print("\nExams with unmet requirements:")
        if unmet_invig_requirements:
            print("Unmet Invigilator Requirements:")
            for exam_name, time_slot in unmet_invig_requirements:
                print(f"Exam: {exam_name}, Time Slot: {time_slot}")
        if unmet_lead_requirements:
            print("Unmet Lead Requirements:")
            for exam_name, time_slot in unmet_lead_requirements:
                print(f"Exam: {exam_name}, Time Slot: {time_slot}")
    else:
        print("\nAll exam requirements have been met.")

def print_unavailable_assignments(unavailable_assignments):
    if unavailable_assignments:
        print("\nInvigilators assigned to unavailable timeslots:")
        for invigilator_name, time_slot, exam_name in unavailable_assignments:
            print(f"Invigilator: {invigilator_name}, Time Slot: {time_slot}, Exam: {exam_name}")
    else:
        print("\nNo invigilators were assigned to unavailable timeslots.")

def print_unassigned_invigilators(invigilators, results, max_time_slot):
    unassigned_invigilators = [invigilator.name for invigilator in invigilators if not any(results[invigilator.name][slot] for slot in range(1, max_time_slot + 1))]
    if unassigned_invigilators:
        print("\nInvigilators with no assignments:")
        for invigilator_name in sorted(unassigned_invigilators):
            print(f"Invigilator: {invigilator_name}")
    else:
        print("\nAll invigilators have assignments.")

# Main Execution
def main():
    invig_file = get_resource_path('large_invigilators.xlsx')
    exams_file = get_resource_path('large_exam_venues.xlsx')

    exam_sessions, invigilators = import_files(invig_file, exams_file)

    max_time_slot = max(exam_sessions.keys())
    random.shuffle(invigilators)
    x, y, unmet_invig, unmet_lead = create_decision_variables(invigilators, exam_sessions)

    prob = pulp.LpProblem("Invigilator_Allocation", pulp.LpMinimize)

    penalty = create_penalty_matrix(invigilators, exam_sessions)

    prob = define_constraints(prob, invigilators, exam_sessions, x, y, unmet_invig, unmet_lead, max_time_slot)

    prob = define_objective(prob, invigilators, exam_sessions, x, y, unmet_invig, unmet_lead, penalty)

    status, objective_value = solve_problem(prob)

    print(f"Status: {status}")
    print(f"Total Penalty: {objective_value}")

    results = initialize_results(invigilators, max_time_slot)

    used_invigilators = set()
    results, used_invigilators, unavailable_assignments = fill_results(results, invigilators, exam_sessions, x, used_invigilators)

    unmet_invig_requirements, unmet_lead_requirements = identify_unmet_requirements(exam_sessions, unmet_invig, unmet_lead)

    unassigned_invigilators = [invigilator.name for invigilator in invigilators if invigilator.name not in used_invigilators]

    exam_ids = [exam.id for time_slot in exam_sessions for exam in exam_sessions[time_slot]]
    exam_colours = generate_colour_palette(len(exam_ids) + 10)

    exam_colours_mapping = {}
    for i in range(len(exam_ids) + 1):
        colour_index = (i * 6) % len(exam_colours)
        exam_colours_mapping[i] = exam_colours[colour_index]

    export_results_to_excel(invigilators, exam_sessions, results, exam_colours_mapping, unmet_invig_requirements, unmet_lead_requirements, unavailable_assignments, unassigned_invigilators)

    print_unmet_requirements(unmet_invig_requirements, unmet_lead_requirements)
    print_unassigned_invigilators(invigilators, results, max_time_slot)
    print_unavailable_assignments(unavailable_assignments)

if __name__ == "__main__":
    main()
